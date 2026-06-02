from __future__ import annotations

import csv
import json
import re
import uuid
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any


@dataclass
class CleanRecord:
    student_no: str
    student_name: str
    course_name: str
    course_code: str | None
    score: int | float | None
    score_raw: str
    score_type: str
    source_file: str
    source_sheet: str
    source_row: int
    confidence: float = 0.9
    warnings: list[str] = field(default_factory=list)


@dataclass
class SheetRows:
    name: str
    rows: list[list[Any]]


CSV_FIELDS = [
    "student_no",
    "student_name",
    "course_id",
    "course_code",
    "course_name",
    "score",
    "score_raw",
    "score_type",
    "source_file",
    "source_sheet",
    "source_row",
    "confidence",
    "warnings",
]


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm(value: Any) -> str:
    return re.sub(r"[\s\u3000_\-—:：/\\()\[\]（）【】{}<>《》,.，。;；]+", "", text(value).lower())


def similar_course(left: str, right: str) -> bool:
    a = norm(left).replace("课程", "").replace("成绩", "")
    b = norm(right).replace("课程", "").replace("成绩", "")
    return bool(a and b and (a == b or a in b or b in a))


def read_excel_any(file_path: str | Path) -> list[SheetRows]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(path, data_only=True)
        sheets: list[SheetRows] = []
        for ws in workbook.worksheets:
            rows = [
                [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                for r in range(1, ws.max_row + 1)
            ]
            sheets.append(SheetRows(ws.title, rows))
        return sheets

    if suffix == ".xls":
        import pandas as pd

        workbook = pd.read_excel(path, sheet_name=None, header=None, engine="xlrd")
        return [SheetRows(str(name), _pandas_rows(frame)) for name, frame in workbook.items()]

    if suffix == ".xlsb":
        from pyxlsb import open_workbook

        sheets = []
        with open_workbook(str(path)) as workbook:
            for sheet_name in workbook.sheets:
                with workbook.get_sheet(sheet_name) as sheet:
                    rows = [[cell.v for cell in row] for row in sheet.rows()]
                sheets.append(SheetRows(sheet_name, rows))
        return sheets

    raise ValueError(f"暂不支持的文件类型：{suffix}")


def _pandas_rows(frame: Any) -> list[list[Any]]:
    import pandas as pd

    rows: list[list[Any]] = []
    for row in frame.values.tolist():
        rows.append([None if pd.isna(value) else value for value in row])
    return rows


def non_empty_count(row: list[Any]) -> int:
    return sum(1 for value in row if text(value))


def is_note_or_summary_row(row: list[Any]) -> bool:
    joined = "".join(text(value) for value in row)
    first = next((text(value) for value in row if text(value)), "")
    keywords = ("平均", "最高", "最低", "合计", "总人数", "及格率", "说明", "备注")
    return any(key in first or key in joined for key in keywords)


def header_score(row: list[Any]) -> int:
    joined = norm(" ".join(text(value) for value in row))
    score = 0
    for keys, weight in [
        (("学号", "学生编号", "studentno", "studentid", "学籍号"), 5),
        (("姓名", "学生姓名", "studentname", "name"), 4),
        (("课程名", "课程名称", "课程", "科目", "coursename"), 4),
        (("成绩", "总分", "总评", "最终", "综合", "分数", "score"), 4),
        (("课程号", "课程代码", "coursecode"), 2),
    ]:
        if any(norm(key) in joined for key in keys):
            score += weight
    if non_empty_count(row) >= 3:
        score += 1
    return score


def detect_header(rows: list[list[Any]]) -> tuple[int, int]:
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows[:40]):
        score = header_score(row)
        if score > best_score:
            best_idx = idx
            best_score = score
    data_start = best_idx + 1
    # Skip secondary explanation rows such as "方式/属性/学期/性质".
    while data_start < len(rows):
        row = rows[data_start]
        joined = "".join(text(value) for value in row)
        if non_empty_count(row) <= 5 and not any(re.fullmatch(r"\d{4,}", text(value)) for value in row):
            if any(key in joined for key in ("方式", "属性", "学期", "性质")):
                data_start += 1
                continue
        break
    return best_idx, data_start


def map_columns(header: list[Any]) -> dict[str, int | None]:
    mapping: dict[str, int | None] = {
        "student_no": None,
        "student_name": None,
        "course_code": None,
        "course_name": None,
        "score": None,
    }
    final_score_candidates: list[tuple[int, int]] = []
    for idx, value in enumerate(header):
        n = norm(value)
        raw = text(value)
        if not n:
            continue
        if mapping["student_no"] is None and any(k in n for k in ("学号", "学生编号", "studentno", "studentid", "学籍号")):
            mapping["student_no"] = idx
        elif mapping["student_name"] is None and any(k in n for k in ("姓名", "学生姓名", "studentname")):
            mapping["student_name"] = idx
        elif mapping["course_code"] is None and any(k in n for k in ("课程号", "课程代码", "coursecode")):
            mapping["course_code"] = idx
        elif mapping["course_name"] is None and any(k in n for k in ("课程名", "课程名称", "课程", "科目", "coursename")):
            mapping["course_name"] = idx

        priority = None
        if any(k in n for k in ("总评成绩", "最终成绩", "综合成绩")):
            priority = 1
        elif any(k in n for k in ("总分", "总成绩")):
            priority = 2
        elif n in {"成绩", "分数", "得分", "score"}:
            priority = 3
        elif any(k in n for k in ("期末成绩", "卷面成绩", "平时成绩")):
            priority = 4
        if priority is not None:
            final_score_candidates.append((priority, idx))

    if final_score_candidates:
        mapping["score"] = sorted(final_score_candidates)[0][1]
    return mapping


def parse_score(value: Any) -> tuple[int | float | None, str, list[str]]:
    raw = text(value)
    warnings: list[str] = []
    if not raw:
        return None, raw, ["missing_score"]
    try:
        number = float(raw.replace(",", ""))
        score: int | float = int(number) if number.is_integer() else number
        if score < 0 or score > 100:
            warnings.append("score_out_of_range")
        return score, raw, warnings
    except ValueError:
        warnings.append("non_numeric_score")
        if any(word in raw for word in ("缺考", "缓考", "作弊", "免修", "通过", "不通过")):
            warnings.append("special_score_status")
        return None, raw, warnings


def candidate_course_score_columns(header: list[Any], mapping: dict[str, int | None]) -> list[int]:
    used = {idx for idx in mapping.values() if idx is not None}
    candidates = []
    for idx, value in enumerate(header):
        if idx in used:
            continue
        label = text(value)
        if not label or label.lower().startswith("column_"):
            continue
        if any(key in label for key in ("班级", "性别", "备注", "序号", "学院", "专业")):
            continue
        candidates.append(idx)
    return candidates


def build_record(
    *,
    row: list[Any],
    source_file: str,
    source_sheet: str,
    source_row: int,
    student_no_col: int,
    student_name_col: int,
    course_name: str,
    course_code_col: int | None,
    score_col: int,
) -> CleanRecord | None:
    if is_note_or_summary_row(row):
        return None
    student_no = text(row[student_no_col] if student_no_col < len(row) else None)
    if not student_no:
        return None
    student_name = text(row[student_name_col] if student_name_col < len(row) else None)
    course_code = text(row[course_code_col] if course_code_col is not None and course_code_col < len(row) else None) or None
    score, score_raw, warnings = parse_score(row[score_col] if score_col < len(row) else None)
    return CleanRecord(
        student_no=student_no,
        student_name=student_name,
        course_name=course_name,
        course_code=course_code,
        score=score,
        score_raw=score_raw,
        score_type="final",
        source_file=source_file,
        source_sheet=source_sheet,
        source_row=source_row,
        warnings=warnings,
    )


def clean_excel(
    *,
    file_path: str | Path,
    selected_course_name: str,
    selected_course_id: str | None = None,
    import_mode: str = "single_course",
    llm_plan: dict[str, Any] | None = None,
    output_dir: str | Path = "outputs",
) -> dict[str, Any]:
    path = Path(file_path)
    warnings: list[str] = []
    errors: list[str] = []
    review_items: list[dict[str, Any]] = []
    records: list[CleanRecord] = []
    field_mapping: dict[str, str] = {}
    table_type = "unknown"
    original_rows = 0
    skipped_rows = 0

    try:
        sheets = read_excel_any(path)
    except Exception as exc:
        errors.append(str(exc))
        return result_dict(path, import_mode, selected_course_name, selected_course_id, "failed", table_type, [], field_mapping, errors, warnings, review_items, 0, 0, True, output_dir)

    for sheet in sheets:
        rows = [row for row in sheet.rows if any(text(value) for value in row)]
        if not rows:
            continue
        header_idx, data_start = detect_header(rows)
        header = rows[header_idx]
        mapping = map_columns(header)
        for key, idx in mapping.items():
            if idx is not None and idx < len(header):
                field_mapping[text(header[idx]) or f"col_{idx + 1}"] = key

        student_no_col = mapping["student_no"]
        student_name_col = mapping["student_name"]
        course_col = mapping["course_name"]
        score_col = mapping["score"]
        if student_no_col is None or student_name_col is None or score_col is None:
            review_items.append(
                {
                    "type": "missing_required_field",
                    "message": f"表格解析失败：缺少必要列（学号列:{student_no_col}, 姓名列:{student_name_col}, 成绩列:{score_col}）",
                    "sheet": sheet.name,
                }
            )
            continue

        table_type = "long_table" if course_col is not None else "single_course"
        if course_col is None:
            candidates = candidate_course_score_columns(header, mapping)
            if len(candidates) >= 2:
                table_type = "multi_course_wide"
                if import_mode == "single_course":
                    matched = [idx for idx in candidates if similar_course(text(header[idx]), selected_course_name)]
                    if len(matched) != 1:
                        review_items.append(
                            {
                                "type": "course_column_not_found",
                                "message": f"多课程宽表中无法唯一定位课程列：{selected_course_name}",
                                "candidates": [text(header[idx]) for idx in candidates],
                                "sheet": sheet.name,
                            }
                        )
                        continue
                    score_targets = [(selected_course_name, matched[0])]
                else:
                    score_targets = [(text(header[idx]), idx) for idx in candidates]
            else:
                score_targets = [(selected_course_name, score_col)]
        else:
            score_targets = [("", score_col)]

        original_rows += max(0, len(rows) - data_start)
        for row_offset in range(data_start, len(rows)):
            row = rows[row_offset]
            if is_note_or_summary_row(row):
                skipped_rows += 1
                continue
            if course_col is not None:
                row_course = text(row[course_col] if course_col < len(row) else None)
                if import_mode == "single_course" and selected_course_name and not similar_course(row_course, selected_course_name):
                    skipped_rows += 1
                    continue
                score_targets = [(row_course or selected_course_name, score_col)]

            for course_name, target_score_col in score_targets:
                rec = build_record(
                    row=row,
                    source_file=path.name,
                    source_sheet=sheet.name,
                    source_row=row_offset + 1,
                    student_no_col=student_no_col,
                    student_name_col=student_name_col,
                    course_name=course_name or selected_course_name,
                    course_code_col=mapping["course_code"],
                    score_col=target_score_col,
                )
                if rec is None:
                    skipped_rows += 1
                    continue
                records.append(rec)

    seen: set[tuple[str, str]] = set()
    for rec in records:
        key = (rec.student_no, rec.course_name)
        if key in seen:
            rec.warnings.append("duplicate_score")
            review_items.append({"type": "duplicate_score", "message": f"{rec.student_no}/{rec.course_name} 重复"})
        seen.add(key)
    if not records and not errors:
        review_items.append(
            {
                "type": "no_records_extracted",
                "message": "没有抽取到成绩记录，请检查课程名、导入模式或表格字段识别结果。",
            }
        )

    status = "success" if records and not review_items and not errors else "need_review"
    if errors:
        status = "failed"
    manual_review = status != "success"
    result = result_dict(
        path,
        import_mode,
        selected_course_name,
        selected_course_id,
        status,
        table_type,
        records,
        field_mapping,
        errors,
        warnings,
        review_items,
        len(records),
        skipped_rows,
        manual_review,
        output_dir,
    )
    if llm_plan:
        result["llm_plan"] = llm_plan
    export_result(result, output_dir)
    return result


def result_dict(
    path: Path,
    import_mode: str,
    selected_course_name: str,
    selected_course_id: str | None,
    status: str,
    table_type: str,
    records: list[CleanRecord],
    field_mapping: dict[str, str],
    errors: list[str],
    warnings: list[str],
    review_items: list[dict[str, Any]],
    valid_rows: int,
    skipped_rows: int,
    manual_review: bool,
    output_dir: str | Path,
) -> dict[str, Any]:
    return {
        "task_id": str(uuid.uuid4()),
        "file_name": path.name,
        "import_mode": import_mode,
        "selected_course": {"course_id": selected_course_id, "course_name": selected_course_name},
        "status": status,
        "table_type": table_type,
        "course_resolution": {
            "resolved_by": "frontend_context",
            "course_name": selected_course_name,
            "course_id": selected_course_id,
            "confidence": 1.0 if selected_course_name else 0.0,
            "evidence": ["selected_course_name"] if selected_course_name else [],
        },
        "records": [asdict(record) for record in records],
        "field_mapping": field_mapping,
        "errors": errors,
        "warnings": warnings,
        "review_items": review_items,
        "cleaning_report": {
            "original_rows": valid_rows + skipped_rows,
            "valid_rows": valid_rows,
            "skipped_rows": skipped_rows,
            "manual_review_required": manual_review,
        },
        "exported_files": {
            "json": str(Path(output_dir) / "cleaning_result.json"),
            "csv": str(Path(output_dir) / "cleaning_result.csv"),
        },
    }


def export_result(result: dict[str, Any], output_dir: str | Path) -> None:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    json_path = out / "cleaning_result.json"
    csv_path = out / "cleaning_result.csv"
    json_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    with csv_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for rec in result.get("records", []):
            writer.writerow(
                {
                    "student_no": rec.get("student_no", ""),
                    "student_name": rec.get("student_name", ""),
                    "course_id": result.get("selected_course", {}).get("course_id") or "",
                    "course_code": rec.get("course_code", "") or "",
                    "course_name": rec.get("course_name", ""),
                    "score": rec.get("score", ""),
                    "score_raw": rec.get("score_raw", ""),
                    "score_type": rec.get("score_type", ""),
                    "source_file": rec.get("source_file", ""),
                    "source_sheet": rec.get("source_sheet", ""),
                    "source_row": rec.get("source_row", ""),
                    "confidence": rec.get("confidence", ""),
                    "warnings": ";".join(rec.get("warnings", [])),
                }
            )


def snapshot_excel(file_path: str | Path, max_rows: int = 30, max_cols: int = 20) -> dict[str, Any]:
    sheets = read_excel_any(file_path)
    return {
        "file_name": Path(file_path).name,
        "sheets": [
            {
                "sheet_name": sheet.name,
                "sample_rows": [
                    {"row": idx + 1, "values": [text(value) for value in row[:max_cols]]}
                    for idx, row in enumerate(sheet.rows[:max_rows])
                ],
            }
            for sheet in sheets
        ],
    }
